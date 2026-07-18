"""Admin moliyaviy hisobot sahifasi (HTML + XLSX eksport)."""
from datetime import datetime

from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse

from django.shortcuts import render

from .reports import compute_financial_report, resolve_period


def _parse_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


@staff_member_required
def financial_report_view(request):
    period = request.GET.get("period", "this_month")
    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))
    start, end, label = resolve_period(period, start, end)

    report = compute_financial_report(start, end)

    if request.GET.get("export") == "xlsx":
        return _export_xlsx(report, label)

    context = {
        **admin.site.each_context(request),
        "title": "Moliyaviy hisobot",
        "report": report,
        "period": period,
        "period_label": label,
    }
    return render(request, "admin/financial_report.html", context)


def _export_xlsx(report, label):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FCE7F3")  # pink-100
    money_fmt = "#,##0 \"so'm\""

    ws["A1"] = "Ziyora — Moliyaviy hisobot"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Davr: {label}"
    ws["A2"].font = Font(italic=True)

    rows = [
        ("Ko'rsatkich", "Qiymat", None),
        ("Buyurtmalar soni", report["orders_count"], "int"),
        ("Umumiy tushum (yetkazish bilan)", report["revenue"], "money"),
        ("Mahsulot tushumi", report["product_revenue"], "money"),
        ("Tannarx (COGS)", report["cogs"], "money"),
        ("Yalpi foyda", report["gross_profit"], "money"),
        ("Foyda ulushi (margin)", f"{report['margin']}%", None),
        ("Yetkazish yig'imi", report["delivery_fees"], "money"),
        ("— To'lov usuli —", "", None),
        ("Naqd (soni / summa)", report["cash"]["count"], "int"),
        ("Naqd summa", report["cash"]["revenue"], "money"),
        ("O'tkazma (soni)", report["transfer"]["count"], "int"),
        ("O'tkazma summa", report["transfer"]["revenue"], "money"),
        ("— To'lov holati —", "", None),
        ("To'langan (soni)", report["paid"]["count"], "int"),
        ("To'langan summa", report["paid"]["revenue"], "money"),
        ("To'lanmagan (soni)", report["unpaid"]["count"], "int"),
        ("To'lanmagan summa", report["unpaid"]["revenue"], "money"),
        ("— Bekor qilingan —", "", None),
        ("Bekor qilingan (soni)", report["cancelled"]["count"], "int"),
        ("Bekor qilingan qiymat", report["cancelled"]["value"], "money"),
    ]

    start_row = 4
    for i, (label_text, value, kind) in enumerate(rows):
        r = start_row + i
        c_label = ws.cell(row=r, column=1, value=label_text)
        c_value = ws.cell(row=r, column=2, value=value)
        if i == 0:
            c_label.font = bold
            c_value.font = bold
            c_label.fill = header_fill
            c_value.fill = header_fill
        elif str(label_text).startswith("—"):
            c_label.font = Font(bold=True, italic=True)
        if kind == "money":
            c_value.number_format = money_fmt
        c_value.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20

    filename = f"ziyora-hisobot-{report['start']}_{report['end']}.xlsx"
    resp = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp
